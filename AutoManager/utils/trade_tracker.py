# AutoManager/utils/trade_tracker.py

import openpyxl
import os
from datetime import datetime, timedelta
import settings_manager as mgr


def load_recent_trades(minutes: int = 720) -> list[dict]:
    """
    Carica i trade chiusi dal file Excel `TRADE_HISTORY_FILE` entro gli ultimi `minutes`.
    Ritorna lista con dizionari nel formato richiesto da update_bot_stats:
        [{ "timestamp": datetime, "profit": float, "correct": bool }, ...]
    """
    file_path = mgr.TRADE_HISTORY_FILE
    if not os.path.exists(file_path):
        return []

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Trades" not in wb.sheetnames:
        return []

    ws = wb["Trades"]
    trades = []
    now = datetime.utcnow()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue  # skip righe vuote o senza timestamp

        try:
            ts = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
            if (now - ts) > timedelta(minutes=minutes):
                continue

            profit = float(row[8]) if row[8] is not None else 0.0
            outcome = str(row[13]).lower()
            correct = outcome == "win"

            trades.append({
                "timestamp": ts,
                "profit": profit,
                "correct": correct
            })

        except Exception:
            continue  # ignora righe con formato errato

    return trades