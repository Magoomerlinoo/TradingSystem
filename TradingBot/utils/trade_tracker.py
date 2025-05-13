# TradingBot/utils/trade_tracker.py

import openpyxl
from openpyxl.utils import get_column_letter
import os
import settings
from datetime import datetime

def initialize_trade_history():
    file_path = settings.TRADE_HISTORY_FILE
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        headers = [
            "Timestamp", "Ticket", "Symbol", "Direction", "Size", "Entry",
            "SL", "TP", "Profit(%)", "Phase", "Used Indicators", "Confidence",
            "Slippage", "Outcome"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(file_path)

def log_closed_trade(trade, pl_pct):
    file_path = settings.TRADE_HISTORY_FILE
    if not os.path.exists(file_path):
        initialize_trade_history()

    wb = openpyxl.load_workbook(file_path)
    ws = wb["Trades"]

    next_row = ws.max_row + 1
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    requested_price = trade.get("entry")
    filled_price = trade.get("filled_price", requested_price)
    slippage = round(filled_price - requested_price, 5)

    if pl_pct > 0.05:
        outcome = "win"
    elif pl_pct < -0.05:
        outcome = "loss"
    else:
        outcome = "BE"

    row_data = [
        now_str,
        trade.get('ticket'),
        trade.get('symbol'),
        trade.get('direction'),
        trade.get('size'),
        trade.get('entry'),
        trade.get('sl'),
        trade.get('tp'),
        f"{pl_pct:.2f}",
        trade.get('phase'),
        ",".join(trade.get('used_indicators', [])),
        f"{trade.get('confidence', 0.0):.2f}",
        slippage,
        outcome
    ]

    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(file_path)

def log_open_trade(trade):
    file_path = settings.TRADE_HISTORY_FILE
    if not os.path.exists(file_path):
        initialize_trade_history()

    wb = openpyxl.load_workbook(file_path)
    ws = wb["Trades"]
    next_row = ws.max_row + 1
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row_data = [
        now_str,
        trade.get('ticket'),
        trade.get('symbol'),
        trade.get('direction'),
        trade.get('size'),
        trade.get('entry'),
        trade.get('sl'),
        trade.get('tp'),
        "N/A",
        trade.get('phase'),
        ",".join(trade.get('used_indicators', [])),
        f"{trade.get('confidence',0.0):.2f}"
    ]
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(file_path)

def load_recent_trades(minutes: int = settings.PERFORMANCE_TIME_DECAY):
    from datetime import datetime, timedelta
    trades = []
    file_path = settings.TRADE_HISTORY_FILE
    if not os.path.exists(file_path):
        return trades

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Trades"]
    now = datetime.utcnow()

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            timestamp_str = row[0]
            profit_pct = float(row[8])
            outcome = row[13]
            ts = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
            if (now - ts) > timedelta(minutes=minutes):
                continue
            correct = outcome == "win"
            trades.append({"timestamp": ts, "profit": profit_pct, "correct": correct})
        except:
            continue
    return trades